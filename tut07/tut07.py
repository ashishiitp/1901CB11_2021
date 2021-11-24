import csv
from openpyxl import Workbook

def feedback_not_submitted():

    
    ltp_mapping_feedback_type = {1: 'lecture', 2: 'tutorial', 3:'practical'}
    output_file_name = "course_feedback_remaining.xlsx" 
    numb=-1
    x=1
    file_students_registered= open('course_registered_by_all_students.csv','r')
    with file_students_registered:
        reader1=list(csv.reader(file_students_registered))
        with open('studentinfo.csv','r') as f:
            reader4=list(csv.reader(f))
            with open('course_feedback_submitted_by_students.csv','r') as file_feedback_submitted:
                reader3=list(csv.reader(file_feedback_submitted))
                with open('course_master_dont_open_in_excel.csv','r') as file_course_master:
                    reader2=list(csv.reader(file_course_master))
                    wb=Workbook()
                    sheet=wb.active
                    sheet.cell(row=1,column=1).value='roll no'
                    sheet.cell(row=1,column=2).value='register_sem'
                    sheet.cell(row=1,column=3).value='schedule_sem'
                    sheet.cell(row=1,column=4).value='sub no'
                    sheet.cell(row=1,column=5).value='Name'
                    sheet.cell(row=1,column=6).value='email'
                    sheet.cell(row=1,column=7).value='aemail'
                    sheet.cell(row=1,column=8).value='contact'
                    # x=1
                    for row1 in reader1:
                        c=0
                        d=0
                        numb=numb+1
                        if numb==0:
                            continue

                        for row2 in reader2:
                            if row1[3]==row2[0]:
                                ltp_there=row2[2]
                                separate_ltp=ltp_there.split('-')
                                #print(separate_ltp)
                                if int(separate_ltp[0])>0:
                                    c=c+1
                                if int(separate_ltp[1])>0:
                                    c=c+1
                                if int(separate_ltp[2])>0:
                                    c=c+1

                        feedback_type = set()  
                        for row3 in reader3:
                            if row1[0]==row3[3] and row1[3]==row3[4]:
                                feedback_type.add(row3[5])
                        if len(feedback_type)!=c:
                            data = list()
                            data.append(row1[0])
                            data.append(row1[1])
                            data.append(row1[2])
                            data.append(row1[3])
                            data.append("NA_IN_STUDENTINFO")
                            data.append("NA_IN_STUDENTINFO")
                            data.append("NA_IN_STUDENTINFO")
                            data.append("NA_IN_STUDENTINFO")
                            
                            for row4 in reader4:
                                if row1[0]==row4[1]:
                                    data[4]=row4[0] 
                                    data[5]=row4[8]
                                    data[6]=row4[9]
                                    data[7]=row4[10]
                            x=x+1
                            sheet.append(data)
                    wb.save('course_feedback_remaining.xlsx')     

                                 

feedback_not_submitted()
 
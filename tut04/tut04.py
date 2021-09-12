import csv
from openpyxl import Workbook
def output_by_subject():
   f= open('regtable_old.csv','r')
   with f:
       reader1=csv.reader(f)
       for row1 in reader1:
          if row1[0]=='rollno':
               continue
          wb=Workbook()
          #wb=load_workbook(r'output_by_subject/'+row[3]+'.xlsx') 
          
          with open ('regtable_old.csv','r') as file: 
              reader2=csv.reader(file)
              #for sheet in range(reader2):
              sheet=wb.active
              sheet.cell(row=1 ,column=1).value = 'rollno'
              sheet.cell(row=1 ,column=2).value = 'register_sum'
              sheet.cell(row=1 ,column=3).value = 'subno'
              sheet.cell(row=1 ,column=4).value = 'sub_type'   
              j=1
              for val in reader2:
                  if val[3]==row1[3]:
                        j+=1
                        sheet.cell(row=j, column=1).value = val[0]
                        sheet.cell(row=j, column=2).value = val[1]
                        sheet.cell(row=j, column=3).value = val[3]
                        sheet.cell(row=j, column=4).value = val[8]
              wb.save(r'output_by_subject/'+row1[3]+'.xlsx')
                        
   return 

def output_individual_roll():
    f= open('regtable_old.csv','r')
    with f:
       reader1=csv.reader(f)
       for row1 in reader1:
          if row1[0]=='rollno':
               continue
          wb=Workbook()
          #wb=load_workbook(r'output_by_subject/'+row[3]+'.xlsx') 
          
          with open ('regtable_old.csv','r') as file: 
              reader2=csv.reader(file)
              #for sheet in range(reader2):
              sheet=wb.active
              sheet.cell(row=1 ,column=1).value = 'rollno'
              sheet.cell(row=1 ,column=2).value = 'register_sum'
              sheet.cell(row=1 ,column=3).value = 'subno'
              sheet.cell(row=1 ,column=4).value = 'sub_type'   
              j=1
              for val in reader2:
                  if val[0]==row1[0]:
                        j+=1
                        sheet.cell(row=j, column=1).value = val[0]
                        sheet.cell(row=j, column=2).value = val[1]
                        sheet.cell(row=j, column=3).value = val[3]
                        sheet.cell(row=j, column=4).value = val[8]
              wb.save(r'output_individual_roll/'+row1[0]+'.xlsx')
    return

output_by_subject()
output_individual_roll()
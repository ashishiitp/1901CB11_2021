import csv
from openpyxl import Workbook
import openpyxl
import os

def filling_grades():
    file1=open("subjects_master.csv",'r')
    with file1:
        massub=csv.reader(file1)             
        codessbname={}                               
        for var in massub:                     
            codessbname[var[0]] = var[1], var[2]       

    file2=open("grades.csv",'r')    
    with file2:             
        rg=csv.reader(file2)                   
        c = 0
        for itr in rg:      
            if c == 0:
                c = 1
                continue
            if os.path.exists(r'output/'+itr[0]+'.xlsx'):                 
                wb =openpyxl.load_workbook(r'output/'+itr[0]+'.xlsx')    
                m = 1
                while m < 9:                                    
                    if(itr[1] == str(m)):                                 
                        sheetname ='sem'+str(m)
                        sheet=wb[sheetname]                        
                        fRows=sheet.max_row+1                                 
                        sheet.cell(row=fRows,column=1).value = fRows      
                        sheet.cell(row=fRows,column=2).value = itr[2]     
                        sheet.cell(row=fRows,column=3).value = codessbname[itr[2]][0]    
                        sheet.cell(row=fRows,column=4).value = codessbname[itr[2]][1]   
                        sheet.cell(row=fRows,column=5).value = int( itr[3] )              
                        sheet.cell(row=fRows,column=6).value = itr[5]                 
                        sheet.cell(row=fRows,column=7).value = itr[4].strip()                    
                    m= m+1                                                               
            else:           
                wb = openpyxl.Workbook() 
                overall = wb.active
                wb.create_sheet( index=0, title="overall")                          
                m = 1
                while m < 9:                                                            
                    wb.create_sheet(index=m,title=r"sem"+str(m))               
                    sheet=wb['sem'+str(m)]                                        
                    sheet.cell(row=1 ,column=1).value = 'Sl No'
                    sheet.cell(row=1 ,column=2).value = 'Subject No'
                    sheet.cell(row=1 ,column=3).value = 'Subject Name'
                    sheet.cell(row=1 ,column=4).value = 'L-T-P'
                    sheet.cell(row=1 ,column=5).value = 'Credit'
                    sheet.cell(row=1 ,column=6).value = 'Sub Type'
                    sheet.cell(row=1 ,column=7).value = 'Grade'
                    m=m+1
            wb.save(r'output/'+itr[0]+'.xlsx')                               
    return


def filling_cpispi():
    f=open("names-roll.csv",'r') 
    with f:
        reader1=csv.reader(f)
        dictstd={}
        for student in reader1:
            dictstd[student[0]] = student[1]
    
    grddict={
        'AA':10,
        'AA*':10,
        'AB':9,
        'AB*':9,
        'BB':8,
        'BB*':8,
        'BC':7,
        'BC*':7,
        'CC':6,
        'CC*':6,
        'CD':5,
        'CD*':5,
        'DD':4,
        'DD*':4,
        'F':0,
        'F*':0,
        'I':0,
        'I*':0
    }
    f2=open("grades.csv",'r')
    with f2:
         # open grades.csv
        y=csv.reader(f2) 
        c=0
        for idx in y:      
            if c==0:
                c=1
                continue
            wb =openpyxl.load_workbook(r'output/'+idx[0]+'.xlsx')
            sheet=wb['overall']
            sheet.cell(row=1,column=1).value ='Roll No'
            sheet.cell(row=1,column=2).value = idx[0]
            
            sheet.cell(row=2,column=1).value ='Name of Student'
            sheet.cell(row=2,column=2).value = dictstd[str(idx[0])]
            
            sheet.cell(row=3,column=1).value ='Discipline'
            sheet.cell(row=3,column=2).value = idx[0][4:6]
            
            sheet.cell(row=4,column=1).value ='Semester Number'
            for y in range(1,9):
                sheet.cell(row = 4, column = y+1).value = y
        
            sheet.cell(row=5,column=1).value ='Semester wise Credit taken'
        
            sheet.cell(row=6,column=1).value ='SPI'
    
            sheet.cell(row=7,column=1).value ='Total Credit Taken'
    
            sheet.cell(row=8,column=1).value ='CPI'

            val=0
            p=0
            q=0
            r=0 
            j=1
            # loop to fill Semester Wise Credit Taken
            while j < 9:
                src=wb['sem'+str(j)]
                val=0 
                insertedinrow=src.max_row+1
                for r1 in range(2,insertedinrow):
                    y = src.cell(row=r1,column=5).value
                    val =val + y
                sheet.cell(row=5,column=j+1).value =val
                j = j+1
            
            
            j=2
            while j < 10:
                y = sheet.cell(row=5,column=j).value
                p=p+y
                sheet.cell(row = 7, column = j).value = p
                j = j+1   
            
            
            j=1
            while j <= 8:
                val=0
                src=wb['sem'+str(j)]
                insertedinrow=src.max_row
                for r1 in range(2,insertedinrow+1):
                    y = src.cell(row=r1,column=5).value * grddict[src.cell(row=r1,column=7).value]
                    val = val+y
                p = sheet.cell(row=5,column=j+1).value
                if(p==0):
                    val = 1
                else:
                    val=val/p 
                sheet.cell(row=6,column=j+1).value =val
                j = j+1

        
            j=2
            p=0
            while j <= 9:
                y = sheet.cell(row=5,column=j).value * sheet.cell(row=6,column=j).value
                p=p+y
                q = sheet.cell(row=7,column=j).value
                if(q > 0 ):
                    r = p / q
                else:
                    r = 1
                sheet.cell(row = 8, column = j).value = r     
                j = j+1
            
            wb.save(r'output/'+idx[0]+'.xlsx')
    return           

def generate_marksheet():
   filling_grades()
   filling_cpispi()

generate_marksheet()